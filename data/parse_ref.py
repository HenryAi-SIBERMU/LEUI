"""
Parse & reshape all ref/datamentah/*.xlsx → data/processed/*.csv
Corrected version: proper CSV names, removed duplicate IKK file.
"""
import pandas as pd
from openpyxl import load_workbook
import os, json
from datetime import datetime

SRC = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ref', 'datamentah')
OUT = os.path.join(os.path.dirname(__file__), 'processed')
os.makedirs(OUT, exist_ok=True)

# Clean old CSVs
for f in os.listdir(OUT):
    if f.endswith('.csv') or f.endswith('.json'):
        os.remove(os.path.join(OUT, f))

metadata = {}


def find_data_start(df_raw):
    for i in range(len(df_raw)):
        val = df_raw.iloc[i, 0]
        if isinstance(val, datetime):
            return i
    return None


def extract_ceic_meta(df_raw, max_rows=25):
    meta = {}
    for i in range(min(max_rows, len(df_raw))):
        key = str(df_raw.iloc[i, 0]) if pd.notna(df_raw.iloc[i, 0]) else ''
        val = df_raw.iloc[i, 1] if df_raw.shape[1] > 1 and pd.notna(df_raw.iloc[i, 1]) else ''
        if key in ('Region', 'Frequency', 'Unit', 'Source', 'First Obs. Date', 'Last Obs. Date'):
            meta[key] = str(val)
    return meta


def melt_investment_data(df_raw, data_start, headers):
    records = []
    for idx in range(data_start, len(df_raw)):
        date_val = df_raw.iloc[idx, 0]
        if not isinstance(date_val, datetime):
            continue
        for col_idx in range(1, min(len(headers), df_raw.shape[1])):
            val = df_raw.iloc[idx, col_idx]
            header = str(headers[col_idx]) if pd.notna(headers[col_idx]) else ''
            if pd.notna(val):
                parts = header.split(':')
                provinsi = parts[2].strip() if len(parts) > 2 else ''
                kabupaten = parts[3].strip() if len(parts) > 3 else ''
                records.append({
                    'date': date_val,
                    'provinsi': provinsi,
                    'kabupaten': kabupaten,
                    'nilai_idr_bn': val
                })
    df = pd.DataFrame(records)
    if len(df) > 0:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['nilai_idr_bn'] = pd.to_numeric(df['nilai_idr_bn'], errors='coerce')
        df = df.dropna(subset=['date', 'nilai_idr_bn'])
        df = df.sort_values(['date', 'provinsi', 'kabupaten']).reset_index(drop=True)
    return df


# ──────────────────────────────────────────────────────────────
# 1. Biaya Investasi (ICOR).xlsx → icor_nasional.csv
# ──────────────────────────────────────────────────────────────
print("=" * 60)
print("1. Biaya Investasi (ICOR).xlsx → icor_nasional.csv")
wb = load_workbook(os.path.join(SRC, 'Biaya Investasi (ICOR).xlsx'), data_only=True)
ws = wb['Chart']
rows = []
for row in ws.iter_rows(values_only=True):
    if row[0] and isinstance(row[0], datetime):
        rows.append({
            'date': row[0],
            'investasi_pmdn': row[1],
            'investasi_pma': row[2],
            'gdp_growth_pct': row[3],
            'icor_pmdn': row[4],
            'icor_pma': row[5] if len(row) > 5 else None
        })
wb.close()
df = pd.DataFrame(rows)
df['date'] = pd.to_datetime(df['date'])
for c in ['investasi_pmdn', 'investasi_pma', 'gdp_growth_pct', 'icor_pmdn', 'icor_pma']:
    df[c] = pd.to_numeric(df[c], errors='coerce')
df = df.sort_values('date').reset_index(drop=True)
out = os.path.join(OUT, 'icor_nasional.csv')
df.to_csv(out, index=False)
print(f"  → icor_nasional.csv | {df.shape} | {df['date'].min()} to {df['date'].max()}")
metadata['icor_nasional'] = {
    'source': 'Biaya Investasi (ICOR).xlsx', 'sheet': 'Chart',
    'rows': len(df), 'cols': list(df.columns),
    'range': [str(df['date'].min()), str(df['date'].max())],
    'unit': 'IDR Billion (investasi), ratio (ICOR)', 'freq': 'Yearly'
}

# ──────────────────────────────────────────────────────────────
# 2. Data Realisasi Investasi.xlsx → 2 files
#    → realisasi_investasi_asing.csv
#    → realisasi_investasi_domestik.csv
# ──────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("2. Data Realisasi Investasi.xlsx")
xl = pd.ExcelFile(os.path.join(SRC, 'Data Realisasi Investasi.xlsx'))

for sn in xl.sheet_names:
    df_raw = pd.read_excel(xl, sheet_name=sn, header=None)

    if 'CAPITAL' in sn.upper() or 'OUTFLOW' in sn.upper():
        print(f"  [{sn}] Capital Outflow (will be parsed from PMI file, skipping)")
        continue

    data_start = find_data_start(df_raw)
    if data_start is None:
        print(f"  [{sn}] No data found, skipping")
        continue

    headers = df_raw.iloc[0].tolist()
    df = melt_investment_data(df_raw, data_start, headers)

    # Determine if Foreign or Domestic from headers
    is_foreign = any('Foreign' in str(h) for h in headers[:10] if pd.notna(h))
    csv_name = 'realisasi_investasi_asing.csv' if is_foreign else 'realisasi_investasi_domestik.csv'
    key_name = csv_name.replace('.csv', '')

    out = os.path.join(OUT, csv_name)
    df.to_csv(out, index=False)
    print(f"  [{sn}] → {csv_name} | {df.shape} | prov={df['provinsi'].nunique()} | kab={df['kabupaten'].nunique()}")

    metadata[key_name] = {
        'source': 'Data Realisasi Investasi.xlsx', 'sheet': sn,
        'rows': len(df), 'unique_provinsi': int(df['provinsi'].nunique()),
        'unique_kabupaten': int(df['kabupaten'].nunique()),
        'range': [str(df['date'].min()), str(df['date'].max())],
        'unit': 'IDR Billion', 'freq': 'Quarterly'
    }

# ──────────────────────────────────────────────────────────────
# 3. Indeks Kepercayaan Konsumen (Expect vs Present).xlsx
#    → ikk_expect_vs_present.csv
# ──────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("3. Indeks Kepercayaan Konsumen (Expect vs Present).xlsx → ikk_expect_vs_present.csv")
df_raw = pd.read_excel(os.path.join(SRC, 'Indeks Kepercayaan Konsumen (Expect vs Present).xlsx'),
                        sheet_name='My Series', header=None)
data_start = find_data_start(df_raw)
if data_start:
    df = df_raw.iloc[data_start:, :3].copy()
    df.columns = ['date', 'ikk_expectation', 'ikk_present']
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    for c in ['ikk_expectation', 'ikk_present']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df['ikk_gap'] = df['ikk_expectation'] - df['ikk_present']
    df = df.dropna(subset=['date']).sort_values('date').reset_index(drop=True)
    out = os.path.join(OUT, 'ikk_expect_vs_present.csv')
    df.to_csv(out, index=False)
    print(f"  → ikk_expect_vs_present.csv | {df.shape} | {df['date'].min()} to {df['date'].max()}")
    metadata['ikk_expect_vs_present'] = {
        'source': 'Indeks Kepercayaan Konsumen (Expect vs Present).xlsx',
        'sheet': 'My Series', 'rows': len(df), 'cols': list(df.columns),
        'range': [str(df['date'].min()), str(df['date'].max())],
        'unit': 'Index', 'freq': 'Monthly'
    }

# ──────────────────────────────────────────────────────────────
# 4. Indeks Kepercayaan Konsumen.xlsx
#    → SKIPPED (duplikat: berisi Investment Realization Foreign,
#      bukan IKK. Sama persis dengan file #2 sheet Foreign)
# ──────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("4. Indeks Kepercayaan Konsumen.xlsx → SKIPPED (duplikat Investment Realization Foreign)")
metadata['_skipped_indeks_kepercayaan_konsumen'] = {
    'source': 'Indeks Kepercayaan Konsumen.xlsx',
    'reason': 'File mislabeled. Contains Investment Realization: Foreign data, identical to Data Realisasi Investasi.xlsx sheet Foreign.',
    'action': 'Skipped — no CSV generated'
}

# ──────────────────────────────────────────────────────────────
# 5. PMI dan Capital Outflow.xlsx
#    → pmi_manufaktur.csv
#    → capital_outflow.csv
# ──────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("5. PMI dan Capital Outflow.xlsx")
xl = pd.ExcelFile(os.path.join(SRC, 'PMI dan Capital Outflow.xlsx'))

for sn in xl.sheet_names:
    df_raw = pd.read_excel(xl, sheet_name=sn, header=None)
    ceic = extract_ceic_meta(df_raw)
    data_start = find_data_start(df_raw)

    if data_start is None:
        print(f"  [{sn}] No data, skipping")
        continue

    df = df_raw.iloc[data_start:, :2].copy()
    df['date'] = pd.to_datetime(df.iloc[:, 0], errors='coerce')
    df['value'] = pd.to_numeric(df.iloc[:, 1], errors='coerce')
    df = df[['date', 'value']].dropna(subset=['date']).sort_values('date').reset_index(drop=True)

    if 'PMI' in sn.upper() or 'MANUFAKTUR' in sn.upper():
        df.columns = ['date', 'pmi_index']
        csv_name = 'pmi_manufaktur.csv'
        metadata['pmi_manufaktur'] = {
            'source': 'PMI dan Capital Outflow.xlsx', 'sheet': sn,
            'rows': len(df), 'cols': list(df.columns),
            'range': [str(df['date'].min()), str(df['date'].max())],
            'unit': 'Index (50=neutral)', 'freq': 'Monthly',
            'data_source': ceic.get('Source', 'S&P Global')
        }
    else:
        df.columns = ['date', 'net_sell_idr_tn']
        csv_name = 'capital_outflow.csv'
        metadata['capital_outflow'] = {
            'source': 'PMI dan Capital Outflow.xlsx', 'sheet': sn,
            'rows': len(df), 'cols': list(df.columns),
            'range': [str(df['date'].min()), str(df['date'].max())],
            'unit': ceic.get('Unit', 'IDR tn'), 'freq': ceic.get('Frequency', 'Daily'),
            'data_source': ceic.get('Source', 'Bank Indonesia')
        }

    out = os.path.join(OUT, csv_name)
    df.to_csv(out, index=False)
    print(f"  [{sn}] → {csv_name} | {df.shape} | {df['date'].min()} to {df['date'].max()}")

# ──────────────────────────────────────────────────────────────
# Save metadata
# ──────────────────────────────────────────────────────────────
meta_path = os.path.join(OUT, 'metadata.json')
with open(meta_path, 'w', encoding='utf-8') as f:
    json.dump(metadata, f, indent=2, ensure_ascii=False, default=str)

# Summary
print(f"\n{'='*60}")
print("SUMMARY — data/processed/")
print(f"{'='*60}")
for f in sorted(os.listdir(OUT)):
    if f.endswith('.csv'):
        fp = os.path.join(OUT, f)
        sz = os.path.getsize(fp)
        nrows = sum(1 for _ in open(fp, encoding='utf-8')) - 1
        print(f"  {f:<50s} {nrows:>7,} rows  ({sz/1024:.0f} KB)")
print(f"  {'metadata.json':<50s}")
print(f"\nTotal CSV files: {len([f for f in os.listdir(OUT) if f.endswith('.csv')])}")
